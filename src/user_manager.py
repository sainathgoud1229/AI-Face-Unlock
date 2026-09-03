import os
import json
import time
import csv
import numpy as np
import config
from supabase_client import supabase


class UserManager:
    def __init__(self):
        self.db_path = config.USERS_DB_PATH
        self.csv_path = os.path.join(config.DATA_DIR, "users_database.csv")
        self.users = {}
        self.load_database()

    # ─────────────────────────────────────────────────────────
    # DATABASE I/O (JSON + CSV AUTO-SYNC)
    # ─────────────────────────────────────────────────────────
    def load_database(self):
        if os.path.exists(self.db_path):
            try:
                with open(self.db_path, "r") as f:
                    self.users = json.load(f)
            except Exception as e:
                print(f"[UserManager] DB load error: {e}")
                self.users = {}
        else:
            self.users = {}

        # On Vercel cold start: pull ALL users from Supabase into memory
        # so face matching works immediately without re-registration
        if not self.users and supabase.enabled:
            print("[UserManager] Cold start detected. Loading users from Supabase...")
            cloud_users = supabase.fetch_all_users()
            for cu in cloud_users:
                uid = cu.get("user_id")
                if uid:
                    cu["feature"] = cu.pop("feature_vector", [])
                    self.users[uid] = cu
            if self.users:
                print(f"[UserManager] Loaded {len(self.users)} user(s) from Supabase cloud.")

        # Migrate legacy face_features.npy
        if not self.users and os.path.exists(config.LEGACY_FEATURE_FILE):
            try:
                feat = np.load(config.LEGACY_FEATURE_FILE)
                self.register_user("Primary User", feat.tolist(), role="Admin")
                print("[UserManager] Migrated legacy face_features.npy")
            except Exception as e:
                print(f"[UserManager] Migration error: {e}")

        # Only export to disk in persistent environments
        if not config.IS_VERCEL:
            self.export_csv()


    def save_database(self):
        if config.IS_VERCEL:
            # On Vercel (stateless/read-only disk), skip file writes.
            # All persistence is handled via Supabase directly.
            return
        try:
            with open(self.db_path, "w") as f:
                json.dump(self.users, f, indent=2)
            self.export_csv()
            self.export_excel()
        except Exception as e:
            print(f"[UserManager] DB save error: {e}")


    def export_excel(self):
        """Export user metadata, face photo, social links, and uploaded PDFs to CSV and native XLSX Excel file."""
        fieldnames = ["Faceid", "name", "Face Image", "Enrolled Date", "github", "Linkedin", "Instagram", "pds", "remaining"]
        rows = []

        for idx, u in enumerate(self.users.values(), start=1):
            shortcuts = u.get("shortcuts", [])
            user_files = u.get("files", [])
            
            github_url = ""
            linkedin_url = ""
            instagram_url = ""
            remaining_list = []
            
            for sc in shortcuts:
                s_name = sc.get("name", "").lower()
                s_url = sc.get("url", "")
                if "github" in s_name or "github.com" in s_url:
                    github_url = s_url
                elif "linkedin" in s_name or "linkedin.com" in s_url:
                    linkedin_url = s_url
                elif "instagram" in s_name or "instagram.com" in s_url:
                    instagram_url = s_url
                else:
                    remaining_list.append(f"{sc.get('name')}: {s_url}")
            
            pdf_list = "; ".join([f.get("name", "") for f in user_files])
            
            rows.append({
                "Faceid": idx,
                "name": u.get("name", ""),
                "Face Image": u.get("face_image", ""),
                "Enrolled Date": u.get("registered_at", ""),
                "github": github_url,
                "Linkedin": linkedin_url,
                "Instagram": instagram_url,
                "pds": pdf_list,
                "remaining": "; ".join(remaining_list),
            })

        # 1. Save CSV
        try:
            with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for r in rows:
                    writer.writerow(r)
        except Exception as e:
            print(f"[UserManager] CSV export error: {e}")

        # 2. Save Native XLSX Excel Spreadsheet
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registered Identities"

            # Styles
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Calibri", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="D1D5DB"),
                right=Side(style="thin", color="D1D5DB"),
                top=Side(style="thin", color="D1D5DB"),
                bottom=Side(style="thin", color="D1D5DB")
            )

            # Write Header
            ws.append(fieldnames)
            for col_num in range(1, len(fieldnames) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            # Write Rows
            for r_idx, r_data in enumerate(rows, start=2):
                row_vals = [r_data[col] for col in fieldnames]
                ws.append(row_vals)
                for c_num in range(1, len(fieldnames) + 1):
                    cell = ws.cell(row=r_idx, column=c_num)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = left_align

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

            xlsx_path = os.path.join(config.DATA_DIR, "users_database.xlsx")
            wb.save(xlsx_path)
        except Exception as e:
            print(f"[UserManager] XLSX export error: {e}")

    def export_csv(self):
        """Export only CSV (called by save_database alongside export_excel)."""
        fieldnames = ["Faceid", "name", "Face Image", "Enrolled Date", "github", "Linkedin", "Instagram", "pds", "remaining"]
        rows = []
        for idx, u in enumerate(self.users.values(), start=1):
            shortcuts = u.get("shortcuts", [])
            user_files = u.get("files", [])
            github_url = ""
            linkedin_url = ""
            instagram_url = ""
            remaining_list = []
            for sc in shortcuts:
                s_name = sc.get("name", "").lower()
                s_url = sc.get("url", "")
                if "github" in s_name or "github.com" in s_url:
                    github_url = s_url
                elif "linkedin" in s_name or "linkedin.com" in s_url:
                    linkedin_url = s_url
                elif "instagram" in s_name or "instagram.com" in s_url:
                    instagram_url = s_url
                else:
                    remaining_list.append(f"{sc.get('name')}: {s_url}")
            pdf_list = "; ".join([f.get("name", "") for f in user_files])
            rows.append({
                "Faceid": idx,
                "name": u.get("name", ""),
                "Face Image": u.get("face_image", ""),
                "Enrolled Date": u.get("registered_at", ""),
                "github": github_url,
                "Linkedin": linkedin_url,
                "Instagram": instagram_url,
                "pds": pdf_list,
                "remaining": "; ".join(remaining_list),
            })
        try:
            with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for r in rows:
                    writer.writerow(r)
        except Exception as e:
            print(f"[UserManager] CSV export error: {e}")


    def add_file(self, user_id, filename, filepath, size_bytes):
        user = self.users.get(user_id)
        if not user:
            return None
        file_id = f"file_{int(time.time())}"
        file_record = {
            "id": file_id,
            "name": filename,
            "path": filepath,
            "size_bytes": size_bytes,
            "size_mb": round(size_bytes / (1024 * 1024), 2),
            "uploaded_at": time.strftime("%Y-%m-%d %H:%M:%S")
        }
        user.setdefault("files", []).append(file_record)
        self.save_database()
        return file_record

    def delete_file(self, user_id, file_id):
        user = self.users.get(user_id)
        if not user:
            return False
        files = user.get("files", [])
        updated = [f for f in files if f["id"] != file_id]
        if len(updated) < len(files):
            user["files"] = updated
            self.save_database()
            return True
        return False

    def get_storage_stats(self, user_id):
        user = self.users.get(user_id)
        if not user:
            return {"used_bytes": 0, "used_mb": 0.0, "total_mb": 1024, "percent": 0.0}
        files = user.get("files", [])
        total_bytes = sum(f.get("size_bytes", 0) for f in files)
        used_mb = round(total_bytes / (1024 * 1024), 2)
        total_mb = 1024.0 # 1GB storage plan
        percent = round(min(100.0, (used_mb / total_mb) * 100), 1)
        return {
            "used_bytes": total_bytes,
            "used_mb": used_mb,
            "total_mb": total_mb,
            "percent": percent
        }


    # ─────────────────────────────────────────────────────────
    # USER CRUD
    # ─────────────────────────────────────────────────────────
    def register_user(self, name, feature_vector, role="User", face_image=None):
        if isinstance(feature_vector, np.ndarray):
            feature_vector = feature_vector.tolist()

        user_id = name.lower().replace(" ", "_")
        existing_shortcuts = self.users.get(user_id, {}).get("shortcuts", [])

        self.users[user_id] = {
            "user_id": user_id,
            "name": name,
            "role": role,
            "registered_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "feature": feature_vector,
            "shortcuts": existing_shortcuts,
            "face_image": face_image,
        }
        self.save_database()
        supabase.sync_user(self.users[user_id])
        return user_id

    def delete_user(self, user_id):
        if user_id in self.users:
            del self.users[user_id]
            self.save_database()
            return True
        return False

    def update_user(self, user_id, name=None, role=None):
        if user_id not in self.users:
            return False
        if name:
            self.users[user_id]["name"] = name
        if role:
            self.users[user_id]["role"] = role
        self.save_database()
        return True

    def list_users(self):
        return [
            {
                "user_id": u["user_id"],
                "name": u["name"],
                "role": u["role"],
                "registered_at": u["registered_at"],
                "face_image": u.get("face_image"),
                "shortcuts_count": len(u.get("shortcuts", [])),
            }
            for u in self.users.values()
        ]

    def get_user(self, user_id):
        if user_id not in self.users and supabase.enabled:
            cloud_user = supabase.fetch_user(user_id)
            if cloud_user:
                cloud_user["feature"] = cloud_user.pop("feature_vector", [])
                self.users[user_id] = cloud_user
        return self.users.get(user_id)

    def match_face(self, query_feature, recognizer, threshold=config.COSINE_SIMILARITY_THRESHOLD):
        if not self.users:
            return None, 0.0

        if isinstance(query_feature, list):
            query_feature = np.array(query_feature, dtype=np.float32)

        best_match = None
        highest_score = -1.0

        for user in self.users.values():
            reg_feat = np.array(user["feature"], dtype=np.float32)
            similarity = recognizer.match(reg_feat, query_feature, 0)
            if similarity > highest_score:
                highest_score = similarity
                if similarity >= threshold:
                    best_match = user

        return best_match, float(highest_score)

    # ─────────────────────────────────────────────────────────
    # PER-USER SHORTCUTS CRUD
    # ─────────────────────────────────────────────────────────
    def get_shortcuts(self, user_id):
        # On Vercel, shortcuts are the source of truth in Supabase
        if config.IS_VERCEL and supabase.enabled:
            cloud_shortcuts = supabase.fetch_shortcuts(user_id)
            if cloud_shortcuts is not None:
                return cloud_shortcuts
        user = self.get_user(user_id)
        if not user:
            return []
        return user.get("shortcuts", [])


    def add_shortcut(self, user_id, name, url, icon="🔗", color="#4facfe", stype="link"):
        user = self.get_user(user_id)
        if not user:
            return None
        if not url.startswith("http"):
            url = "https://" + url
        sid = f"{int(time.time())}_{name.lower().replace(' ', '_')}"
        shortcut = {
            "id": sid,
            "name": name,
            "url": url,
            "icon": icon,
            "color": color,
            "type": stype,
            "added_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        user.setdefault("shortcuts", []).append(shortcut)
        # On Vercel we skip disk writes; sync directly to Supabase
        if config.IS_VERCEL:
            supabase.sync_shortcut(user_id, shortcut)
        else:
            self.save_database()
            supabase.sync_shortcut(user_id, shortcut)
        return shortcut


    def update_shortcut(self, user_id, shortcut_id, name=None, url=None, icon=None, color=None):
        user = self.get_user(user_id)
        if not user:
            return False
        for s in user.get("shortcuts", []):
            if s["id"] == shortcut_id:
                if name: s["name"] = name
                if url: s["url"] = url
                if icon: s["icon"] = icon
                if color: s["color"] = color
                if not config.IS_VERCEL:
                    self.save_database()
                supabase.sync_shortcut(user_id, s)
                return True

        return False

    def delete_shortcut(self, user_id, shortcut_id):
        user = self.get_user(user_id)
        if not user:
            return False
        original = user.get("shortcuts", [])
        updated = [s for s in original if s["id"] != shortcut_id]
        if len(updated) < len(original):
            user["shortcuts"] = updated
            if not config.IS_VERCEL:
                self.save_database()
            # Also delete from Supabase
            supabase.delete_shortcut(shortcut_id)
            return True
        # If not in local memory (Vercel), delete directly from Supabase
        if config.IS_VERCEL:
            return supabase.delete_shortcut(shortcut_id)
        return False


